from openpyxl import load_workbook


def main():
    wb = load_workbook("excel_wb.xlsx")
    print(f"Workbook Sheets: {wb.sheetnames}")
    for sheet in wb.sheetnames:
        print(
            f"{sheet} sheet has {wb[sheet].max_row} rows and {wb[sheet].max_column} columns"
        )

    devices = wb["Devices"]
    device_list = []
    for row in range(2, devices.max_row + 1):
        device_list.append(devices.cell(row=row, column=1).value)
    print(device_list)


if __name__ == "__main__":
    main()
