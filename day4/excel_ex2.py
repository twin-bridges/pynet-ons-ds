from pprint import pprint
from openpyxl import load_workbook


def main():
    wb = load_workbook("excel_wb.xlsx")

    sites = wb["Sites"]
    ip_addresses = {}
    for row in range(2, sites.max_row + 1):
        ip_addresses[sites.cell(row=row, column=1).value] = {}
        ip_addresses[sites.cell(row=row, column=1).value]["devices"] = sites.cell(
            row=row, column=2
        ).value.split(",")

    devices = wb["Devices"]
    for row in range(2, devices.max_row + 1):
        device = devices.cell(row=row, column=1).value
        ip = devices.cell(row=row, column=3).value
        for site, info in ip_addresses.copy().items():
            if device in info["devices"]:
                ip_addresses[site][device] = ip
    for site, info in ip_addresses.copy().items():
        ip_addresses[site].pop("devices")

    pprint(ip_addresses)


if __name__ == "__main__":
    main()
