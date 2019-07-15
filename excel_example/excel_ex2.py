from openpyxl import load_workbook


def main():
    wb = load_workbook("excel_wb.xlsx")
    print(f"Workbook Sheets: {wb.sheetnames}")
    users = wb["Users"]
    users_dict = {}
    for row in range(2, users.max_row + 1):
        username = users.cell(row=row, column=1).value
        users_dict[username] = {}
        users_dict[username]["fname"] = users.cell(row=row, column=2).value
        users_dict[username]["lname"] = users.cell(row=row, column=3).value
        users_dict[username]["role"] = users.cell(row=row, column=4).value
        users_dict[username]["site"] = users.cell(row=row, column=5).value
    print(users_dict)


if __name__ == "__main__":
    main()
