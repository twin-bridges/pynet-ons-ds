from openpyxl import load_workbook


def main():
    wb = load_workbook("excel_wb.xlsx")
    print(f"Workbook Sheets: {wb.sheetnames}")
    users = wb["Users"]
    users_headers = []
    for col in range(1, users.max_column + 1):
        users_headers.append(users.cell(row=1, column=col).value)
    print(users_headers)


if __name__ == "__main__":
    main()
